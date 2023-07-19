from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border

import os

from config import working_path

ind = 11

cells = 'ABCDEFGHIJKLM'

workbook = ''

for file in os.listdir(os.path.join(working_path, f'1583_1')):
    print(file)

    if ind == 11:
        workbook = load_workbook(os.path.join(working_path, f'1583_1\\{file}'))

        sheet1 = workbook.active
        sheet1.unmerge_cells('A12:C12')

    book = load_workbook(os.path.join(working_path, f'1583_1\\{file}'))
    sheet = book.active

    for i in cells:
        try:
            sheet1[i + str(ind)].value = sheet[i + '11'].value
            sheet1[i + str(ind)].alignment = Alignment(
                horizontal=sheet[i + '11'].alignment.horizontal,
                vertical=sheet[i + '11'].alignment.vertical,
                text_rotation=sheet[i + '11'].alignment.text_rotation,
                wrap_text=sheet[i + '11'].alignment.wrap_text,
                shrink_to_fit=sheet[i + '11'].alignment.shrink_to_fit,
                indent=sheet[i + '11'].alignment.indent,
            )
            sheet1[i + str(ind)].fill = PatternFill(
                start_color=sheet[i + '11'].fill.start_color,
                end_color=sheet[i + '11'].fill.end_color,
                fill_type=sheet[i + '11'].fill.fill_type,
            )
            sheet1[i + str(ind)].border = Border(
                left=sheet[i + '11'].border.left,
                right=sheet[i + '11'].border.right,
                top=sheet[i + '11'].border.top,
                bottom=sheet[i + '11'].border.bottom,
            )

        except:
            sheet1.unmerge_cells([f'A{ind}:C{ind}'])
            sheet1[i + str(ind)].value = sheet[i + '11'].value

    ind += 1

workbook.save('result_a.xlsx')