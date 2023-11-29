import datetime

import psycopg2
from openpyxl import load_workbook

from config import db_name, db_username, db_password
































import pandas as pd



def kek(date_, check):

    s = 0

    print(date_, check)

    conn = psycopg2.connect(dbname=db_name, host='172.16.10.22', port='5432',
                            user=db_username, password=db_password)

    cur = conn.cursor(name='1583_first_part')

    query = f"""
           select payload from dwh_stgll."transaction" t
            join dwh_stgll.actor a on a.id = t.actor_id
            join dwh_stgll.card c on c.id = t.card_id
            left join dwh_stgll.clients_loyalty cl on cl.order_hash = t.reference
            left join dwh_data.fact_order_new fon on fon.source_order_id = cl.code_order
            where fon.order_date >= '{date_}' and fon.order_code = '{check}'
       """

    cur.execute(query)

    df = pd.DataFrame(cur.fetchall())

    for rows in df.iterrows():
        # print(rows)
        payment = rows[1].iloc[0]
        print(len(payment.get('payment_types')))
        for pay in payment.get('payment_types'):
            print(pay)
            if pay.get('type') == 'BP':
                s += pay.get('sum')
        # print(payment.get('payment_types'))
    # payment = df['payment_types']
    # print(df)
    # print(a)
    # print(payment[1].get('type'))
    print(s)
    cur.close()
    conn.close()

    return s


if __name__ == '__main__':

    book = load_workbook(r'\\vault.magnum.local\Common\Stuff\_06_Бухгалтерия\Для робота\Процесс Сверка ОПТа\Файл сбора Октябрь 2023.xlsx')

    sheet = book.active
    sum_ = 0
    for row in range(3, sheet.max_row + 1):

        input_date_object = datetime.datetime.strptime(sheet[f'B{row}'].value, "%d.%m.%Y")

        output_date_string = input_date_object.strftime("%Y-%m-%d")

        date_ = output_date_string
        check_id = sheet[f'J{row}'].value
        try:
            ss = kek(date_, check_id)
        except:
            ss = 0

        sum_ += ss

        # break
    print('SUM:', sum_)
